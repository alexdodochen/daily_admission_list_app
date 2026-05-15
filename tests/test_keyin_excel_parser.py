"""Pure-logic tests for keyin_excel_parser internal helpers and
parse_schedule_excel against a synthetic .xlsx file.
"""
from __future__ import annotations

import openpyxl
import pytest

from app.services import keyin_excel_parser as kp


def test_to_int_day_accepts_valid_range():
    assert kp._to_int_day("1") == 1
    assert kp._to_int_day("31") == 31
    assert kp._to_int_day("1.0") == 1


def test_to_int_day_rejects_out_of_range():
    assert kp._to_int_day("0") is None
    assert kp._to_int_day("32") is None
    assert kp._to_int_day("garbage") is None
    assert kp._to_int_day("") is None


def test_normalize_name_strips_whitespace_and_corrects():
    assert kp._normalize_name("廖 瑀") == "廖瑀"
    assert kp._normalize_name("胡晟瀚") == "胡展瀚"  # correction table
    assert kp._normalize_name("黃勝翔") == "黃睦翔"


def test_col_keyword_detects_vs_cr_day():
    assert kp._col_keyword("值班VS") == "vs"
    assert kp._col_keyword("值班CR") == "cr"
    assert kp._col_keyword("日") == "day"
    assert kp._col_keyword("randomstuff") is None


def test_parse_vertical_xlsx(tmp_path):
    """End-to-end: build a vertical-format .xlsx and parse it."""
    p = tmp_path / "vertical.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["日", "值班VS", "值班CR"])
    ws.append([1, "廖瑀", "胡展瀚"])
    ws.append([2, "陳昭佑", "徐麒翔"])
    ws.append([3, "張獻元", ""])
    wb.save(p)

    result = kp.parse_schedule_excel(str(p))
    assert result["ok"] is True
    assert result["format"] == "縱向"
    assert result["vs_schedule"] == {"1": "廖瑀", "2": "陳昭佑", "3": "張獻元"}
    assert result["cr_schedule"] == {"1": "胡展瀚", "2": "徐麒翔"}
    assert result["warnings"] == []


def test_parse_vertical_xlsx_warns_unknown_doctor(tmp_path):
    p = tmp_path / "unknown.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["日", "值班VS"])
    ws.append([1, "陌生人"])
    wb.save(p)

    result = kp.parse_schedule_excel(str(p))
    assert result["ok"] is True
    assert any("陌生人" in w for w in result["warnings"])


def test_parse_unreadable_extension_returns_error(tmp_path):
    p = tmp_path / "blah.csv"
    p.write_text("notexcel")
    result = kp.parse_schedule_excel(str(p))
    assert result["ok"] is False
    assert "無法讀取" in result["error"] or "不支援" in result["error"]


def test_parse_empty_xlsx_returns_error(tmp_path):
    p = tmp_path / "empty.xlsx"
    wb = openpyxl.Workbook()
    wb.save(p)
    # default Workbook() has a single empty active sheet → rows is [[]] or []
    result = kp.parse_schedule_excel(str(p))
    # Either reports empty content or unable to recognise format — both acceptable
    assert result["ok"] is False
